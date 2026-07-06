"""
Output generator module for TallyBridge.

Generates Tally-compatible Excel files from parsed Zerodha Tax P&L data.
Produces separate files for:
  - Profit trades (capital gains income)
  - Loss trades (capital losses)
  - Charges (brokerage, DP charges, etc.)
  - Dividends (dividend income)

Each file uses Tally's expected import format:
  Date | Voucher Type | Particulars | Ledger Name | Debit | Credit | Narration
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.parser import ParsedData, Trade, Charge, Dividend


# Tally ledger name mappings
LEDGER_NAMES = {
    "Equity - Intraday": "Speculative Income",
    "Equity - Short Term": "Short Term Capital Gains",
    "Equity - Long Term": "Long Term Capital Gains",
    "Equity - Buyback": "Income from Buyback",
    "Non Equity": "Non-Equity Capital Gains",
    "Mutual Funds": "Capital Gains - Mutual Funds",
    "F&O": "F&O Trading Income",
    "Currency": "Currency Trading Income",
    "Commodity": "Commodity Trading Income",
}

LOSS_LEDGER_NAMES = {
    "Equity - Intraday": "Speculative Loss",
    "Equity - Short Term": "Short Term Capital Loss",
    "Equity - Long Term": "Long Term Capital Loss",
    "Equity - Buyback": "Loss from Buyback",
    "Non Equity": "Non-Equity Capital Loss",
    "Mutual Funds": "Capital Loss - Mutual Funds",
    "F&O": "F&O Trading Loss",
    "Currency": "Currency Trading Loss",
    "Commodity": "Commodity Trading Loss",
}

# Styling constants
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11, color="555555")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_all(data: ParsedData, output_dir: str) -> list:
    """
    Generate all Tally-compatible Excel files.

    Args:
        data: ParsedData from the parser.
        output_dir: Directory to write output files to.

    Returns:
        List of output file paths created.
    """
    os.makedirs(output_dir, exist_ok=True)
    files = []

    fy_label = _fy_label(data)

    # Generate profit trades file
    if data.profit_trades:
        path = os.path.join(output_dir, f"Profit_Entries_{fy_label}.xlsx")
        _generate_trades_file(data.profit_trades, path, "Profit", data)
        files.append(path)

    # Generate loss trades file
    if data.loss_trades:
        path = os.path.join(output_dir, f"Loss_Entries_{fy_label}.xlsx")
        _generate_trades_file(data.loss_trades, path, "Loss", data)
        files.append(path)

    # Generate charges file
    if data.charges:
        path = os.path.join(output_dir, f"Charges_{fy_label}.xlsx")
        _generate_charges_file(data.charges, path, data)
        files.append(path)

    # Generate dividends file
    if data.dividends:
        path = os.path.join(output_dir, f"Dividends_{fy_label}.xlsx")
        _generate_dividends_file(data.dividends, path, data)
        files.append(path)

    return files


def _fy_label(data: ParsedData) -> str:
    """Generate a financial year label like 'FY2025-26'."""
    if data.financial_year:
        # Parse "2025-04-01 to 2026-03-31"
        parts = data.financial_year.split(" to ")
        if len(parts) == 2:
            start_year = parts[0].split("-")[0]
            end_year = parts[1].split("-")[0][-2:]
            return f"FY{start_year}-{end_year}"
    return "FY_Unknown"


def _generate_trades_file(trades: list, filepath: str, category: str, data: ParsedData):
    """Generate an Excel file for profit or loss trades."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Import"

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"TallyBridge - {category} Entries"
    ws["A1"].font = TITLE_FONT

    # Client info
    ws["A2"] = f"Client: {data.client_name} ({data.client_id}) | PAN: {data.pan}"
    ws["A2"].font = SUBTITLE_FONT

    # Summary
    total = sum(abs(t.profit) for t in trades)
    ws["A3"] = f"Total {category}: ₹{total:,.2f} | Trades: {len(trades)}"
    ws["A3"].font = SUBTITLE_FONT

    # Tally Import Format header (Row 5)
    tally_headers = [
        "Date", "Voucher Type", "Particulars", "Ledger Name",
        "Debit", "Credit", "Narration", "Trade Type",
    ]
    for col, header in enumerate(tally_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    row_num = 6
    for trade in sorted(trades, key=lambda t: (t.trade_type, t.exit_date)):
        is_profit = trade.profit >= 0
        ledger_map = LEDGER_NAMES if is_profit else LOSS_LEDGER_NAMES
        ledger = ledger_map.get(trade.trade_type, "Capital Gains/Losses")

        ws.cell(row=row_num, column=1, value=trade.exit_date)
        ws.cell(row=row_num, column=2, value="Journal")
        ws.cell(row=row_num, column=3, value=f"{trade.symbol} ({trade.isin})")
        ws.cell(row=row_num, column=4, value=ledger)

        if is_profit:
            ws.cell(row=row_num, column=5, value="")  # Debit
            ws.cell(row=row_num, column=6, value=round(trade.profit, 2))  # Credit
        else:
            ws.cell(row=row_num, column=5, value=round(abs(trade.profit), 2))  # Debit
            ws.cell(row=row_num, column=6, value="")  # Credit

        narration = (
            f"Buy: ₹{trade.buy_value:.2f} on {trade.entry_date} | "
            f"Sell: ₹{trade.sell_value:.2f} on {trade.exit_date} | "
            f"Qty: {int(trade.quantity)} | Holding: {trade.period_of_holding}d"
        )
        ws.cell(row=row_num, column=7, value=narration)
        ws.cell(row=row_num, column=8, value=trade.trade_type)

        # Apply border to data cells
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = THIN_BORDER

        row_num += 1

    # Add a detailed sheet with full trade data
    ws_detail = wb.create_sheet("Detailed Trades")
    detail_headers = [
        "Symbol", "ISIN", "Entry Date", "Exit Date", "Quantity",
        "Buy Value", "Sell Value", "Profit", "Trade Type",
        "Period of Holding", "Brokerage", "STT", "Exchange Charges",
        "Stamp Duty", "IGST", "Total Charges",
    ]
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, trade in enumerate(sorted(trades, key=lambda t: (t.trade_type, t.exit_date)), 2):
        ws_detail.cell(row=i, column=1, value=trade.symbol)
        ws_detail.cell(row=i, column=2, value=trade.isin)
        ws_detail.cell(row=i, column=3, value=trade.entry_date)
        ws_detail.cell(row=i, column=4, value=trade.exit_date)
        ws_detail.cell(row=i, column=5, value=trade.quantity)
        ws_detail.cell(row=i, column=6, value=round(trade.buy_value, 2))
        ws_detail.cell(row=i, column=7, value=round(trade.sell_value, 2))
        ws_detail.cell(row=i, column=8, value=round(trade.profit, 2))
        ws_detail.cell(row=i, column=9, value=trade.trade_type)
        ws_detail.cell(row=i, column=10, value=trade.period_of_holding)
        ws_detail.cell(row=i, column=11, value=round(trade.brokerage, 4))
        ws_detail.cell(row=i, column=12, value=round(trade.stt, 4))
        ws_detail.cell(row=i, column=13, value=round(trade.exchange_charges, 4))
        ws_detail.cell(row=i, column=14, value=round(trade.stamp_duty, 4))
        ws_detail.cell(row=i, column=15, value=round(trade.igst, 4))
        ws_detail.cell(row=i, column=16, value=round(trade.total_charges, 4))

    # Auto-fit column widths
    _auto_fit_columns(ws)
    _auto_fit_columns(ws_detail)

    wb.save(filepath)


def _generate_charges_file(charges: list, filepath: str, data: ParsedData):
    """Generate an Excel file for charges/expenses."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Import"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "TallyBridge - Charges & Expenses"
    ws["A1"].font = TITLE_FONT

    ws["A2"] = f"Client: {data.client_name} ({data.client_id}) | PAN: {data.pan}"
    ws["A2"].font = SUBTITLE_FONT

    total_debit = sum(c.debit for c in charges)
    total_credit = sum(c.credit for c in charges)
    ws["A3"] = f"Total Debits: ₹{total_debit:,.2f} | Total Credits: ₹{total_credit:,.2f}"
    ws["A3"].font = SUBTITLE_FONT

    # Headers
    headers = [
        "Date", "Voucher Type", "Particulars", "Ledger Name",
        "Debit", "Credit", "Segment",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data
    row_num = 6
    for charge in sorted(charges, key=lambda c: c.posting_date):
        # Determine ledger name based on the charge type
        ledger = _classify_charge(charge.particulars)

        ws.cell(row=row_num, column=1, value=charge.posting_date)
        ws.cell(row=row_num, column=2, value="Journal")
        ws.cell(row=row_num, column=3, value=charge.particulars)
        ws.cell(row=row_num, column=4, value=ledger)
        ws.cell(row=row_num, column=5, value=round(charge.debit, 2) if charge.debit > 0 else "")
        ws.cell(row=row_num, column=6, value=round(charge.credit, 2) if charge.credit > 0 else "")
        ws.cell(row=row_num, column=7, value=charge.segment)

        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = THIN_BORDER

        row_num += 1

    _auto_fit_columns(ws)
    wb.save(filepath)


def _generate_dividends_file(dividends: list, filepath: str, data: ParsedData):
    """Generate an Excel file for dividend income."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Import"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "TallyBridge - Dividend Income"
    ws["A1"].font = TITLE_FONT

    ws["A2"] = f"Client: {data.client_name} ({data.client_id}) | PAN: {data.pan}"
    ws["A2"].font = SUBTITLE_FONT

    total = sum(d.net_amount for d in dividends)
    ws["A3"] = f"Total Dividend Income: ₹{total:,.2f} | Entries: {len(dividends)}"
    ws["A3"].font = SUBTITLE_FONT

    # Tally format headers
    headers = [
        "Date", "Voucher Type", "Particulars", "Ledger Name",
        "Debit", "Credit", "Narration",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    row_num = 6
    for div in sorted(dividends, key=lambda d: d.ex_date):
        ws.cell(row=row_num, column=1, value=div.ex_date)
        ws.cell(row=row_num, column=2, value="Receipt")
        ws.cell(row=row_num, column=3, value=f"{div.symbol} ({div.isin})")
        ws.cell(row=row_num, column=4, value="Dividend Income")
        ws.cell(row=row_num, column=5, value="")  # Debit
        ws.cell(row=row_num, column=6, value=round(div.net_amount, 2))  # Credit
        narration = f"Dividend from {div.symbol} | Qty: {int(div.quantity)} | ₹{div.dividend_per_share}/share"
        ws.cell(row=row_num, column=7, value=narration)

        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = THIN_BORDER

        row_num += 1

    # Add summary by stock
    ws_summary = wb.create_sheet("Summary by Stock")
    ws_summary.cell(row=1, column=1, value="Symbol").font = HEADER_FONT
    ws_summary.cell(row=1, column=2, value="Total Dividend").font = HEADER_FONT
    ws_summary.cell(row=1, column=3, value="Number of Payments").font = HEADER_FONT

    # Group dividends by symbol
    by_symbol = {}
    for d in dividends:
        if d.symbol not in by_symbol:
            by_symbol[d.symbol] = {"total": 0.0, "count": 0}
        by_symbol[d.symbol]["total"] += d.net_amount
        by_symbol[d.symbol]["count"] += 1

    for i, (symbol, info) in enumerate(
        sorted(by_symbol.items(), key=lambda x: -x[1]["total"]), 2
    ):
        ws_summary.cell(row=i, column=1, value=symbol)
        ws_summary.cell(row=i, column=2, value=round(info["total"], 2))
        ws_summary.cell(row=i, column=3, value=info["count"])

    _auto_fit_columns(ws)
    _auto_fit_columns(ws_summary)
    wb.save(filepath)


def _classify_charge(particulars: str) -> str:
    """Classify a charge into a Tally ledger name based on its description."""
    p = particulars.lower()

    if "dp charge" in p:
        return "DP Charges"
    elif "amc" in p or "demat" in p:
        return "Demat AMC Charges"
    elif "smallcase" in p:
        return "Smallcase Fees"
    elif "brokerage" in p:
        return "Brokerage Expenses"
    elif "gst" in p:
        return "GST on Brokerage"
    elif "stamp" in p:
        return "Stamp Duty"
    elif "stt" in p:
        return "STT Expenses"
    elif "sebi" in p:
        return "SEBI Charges"
    elif "pledge" in p:
        return "Pledge Charges"
    elif "interest" in p:
        return "Interest Charges"
    elif "penalty" in p or "penal" in p:
        return "Penalty Charges"
    elif "call" in p and "trade" in p:
        return "Call & Trade Charges"
    else:
        return "Other Trading Expenses"


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        # Cap at 50 chars wide
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
