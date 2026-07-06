"""
Parser module for Zerodha Tax P&L Excel files.

Extracts trades, charges, and dividends from the standard Zerodha
tax report format (taxpnl-CLIENTID-YYYY_YYYY-Q1-Q4.xlsx).
"""

from dataclasses import dataclass, field
from typing import Optional
import openpyxl


@dataclass
class Trade:
    """A single trade exit entry."""
    symbol: str
    isin: str
    entry_date: str
    exit_date: str
    quantity: float
    buy_value: float
    sell_value: float
    profit: float
    period_of_holding: int
    fair_market_value: float
    taxable_profit: float
    turnover: float
    brokerage: float
    exchange_charges: float
    ipft: float
    sebi_charges: float
    cgst: float
    sgst: float
    igst: float
    stamp_duty: float
    stt: float
    trade_type: str  # "Equity - Intraday", "Equity - Short Term", etc.

    @property
    def total_charges(self) -> float:
        return (
            self.brokerage + self.exchange_charges + self.ipft +
            self.sebi_charges + self.cgst + self.sgst + self.igst +
            self.stamp_duty + self.stt
        )

    @property
    def is_profit(self) -> bool:
        return self.profit >= 0


@dataclass
class Charge:
    """A debit/credit entry from the Other Debits and Credits sheet."""
    particulars: str
    posting_date: str
    debit: float
    credit: float
    segment: str  # "Equity", "Mutual Funds", "F&O", etc.


@dataclass
class Dividend:
    """A dividend income entry."""
    symbol: str
    isin: str
    ex_date: str
    quantity: float
    dividend_per_share: float
    net_amount: float


@dataclass
class ParsedData:
    """Complete parsed data from a Zerodha Tax P&L file."""
    client_id: str = ""
    client_name: str = ""
    pan: str = ""
    financial_year: str = ""
    trades: list = field(default_factory=list)
    charges: list = field(default_factory=list)
    dividends: list = field(default_factory=list)
    total_dividend: float = 0.0

    # Summary from Equity and Non Equity sheet
    intraday_profit: float = 0.0
    short_term_profit: float = 0.0
    long_term_profit: float = 0.0
    non_equity_profit: float = 0.0

    @property
    def profit_trades(self) -> list:
        return [t for t in self.trades if t.is_profit]

    @property
    def loss_trades(self) -> list:
        return [t for t in self.trades if not t.is_profit]


# Known section headers in the Tradewise Exits sheet
TRADE_SECTIONS = [
    "Equity - Intraday",
    "Equity - Short Term",
    "Equity - Long Term",
    "Equity - Buyback",
    "Non Equity",
    "Mutual Funds",
    "F&O",
    "Currency",
    "Commodity",
]

# Expected column headers for trades
TRADE_COLUMNS = [
    "Symbol", "ISIN", "Entry Date", "Exit Date", "Quantity",
    "Buy Value", "Sell Value", "Profit", "Period of Holding",
    "Fair Market Value", "Taxable Profit", "Turnover",
    "Brokerage", "Exchange Transaction Charges", "IPFT",
    "SEBI Charges", "CGST", "SGST", "IGST", "Stamp Duty", "STT",
]


def parse_file(filepath: str) -> ParsedData:
    """
    Parse a Zerodha Tax P&L Excel file.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        ParsedData object with all extracted information.

    Raises:
        ValueError: If the file format is not recognized.
        FileNotFoundError: If the file does not exist.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = ParsedData()

    # Extract client info from any sheet (they all have it in rows 7-9)
    _parse_client_info(wb, data)

    # Parse tradewise exits
    if "Tradewise Exits from 2025-04-01" in wb.sheetnames:
        _parse_tradewise_exits(wb, data)
    else:
        # Find the tradewise exits sheet by prefix
        for name in wb.sheetnames:
            if name.startswith("Tradewise Exits"):
                _parse_tradewise_exits(wb, data, sheet_name=name)
                break

    # Parse summary from Equity and Non Equity sheet
    if "Equity and Non Equity" in wb.sheetnames:
        _parse_equity_summary(wb, data)

    # Parse other debits and credits
    if "Other Debits and Credits" in wb.sheetnames:
        _parse_charges(wb, data)

    # Parse dividends
    if "Equity Dividends" in wb.sheetnames:
        _parse_dividends(wb, data)

    wb.close()
    return data


def _parse_client_info(wb, data: ParsedData):
    """Extract client ID, name, and PAN from the first available sheet."""
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=7, max_row=9, values_only=True):
        if row[1] == "Client ID":
            data.client_id = str(row[2]) if row[2] else ""
        elif row[1] == "Client Name":
            data.client_name = str(row[2]) if row[2] else ""
        elif row[1] == "PAN":
            data.pan = str(row[2]) if row[2] else ""

    # Try to extract FY from any title row
    for row in ws.iter_rows(min_row=11, max_row=11, values_only=True):
        if row[1] and "from" in str(row[1]):
            title = str(row[1])
            # Extract date range like "2025-04-01 to 2026-03-31"
            if "from" in title and "to" in title:
                parts = title.split("from")[-1].strip()
                data.financial_year = parts


def _parse_tradewise_exits(wb, data: ParsedData, sheet_name: Optional[str] = None):
    """Parse the Tradewise Exits sheet for individual trade data."""
    if sheet_name is None:
        # Find sheet by prefix
        for name in wb.sheetnames:
            if name.startswith("Tradewise Exits"):
                sheet_name = name
                break
    if sheet_name is None:
        return

    ws = wb[sheet_name]
    current_section = None
    expect_header = False

    for row in ws.iter_rows(min_row=13, max_row=ws.max_row, values_only=True):
        # Skip empty rows
        if not any(cell is not None for cell in row):
            continue

        cell_b = row[1] if len(row) > 1 else None
        if cell_b is None:
            continue

        cell_b_str = str(cell_b).strip()

        # Check if this is a section header
        if cell_b_str in TRADE_SECTIONS:
            current_section = cell_b_str
            expect_header = True
            continue

        # Skip the column header row
        if cell_b_str == "Symbol":
            expect_header = False
            continue

        # Skip empty string rows
        if cell_b_str == "":
            continue

        # If we have a current section and this isn't a header, it's data
        if current_section and not expect_header:
            try:
                trade = _parse_trade_row(row, current_section)
                if trade:
                    data.trades.append(trade)
            except (ValueError, TypeError, IndexError):
                # Skip malformed rows
                continue


def _parse_trade_row(row: tuple, trade_type: str) -> Optional[Trade]:
    """Parse a single trade data row into a Trade object."""
    # row[0] is always None (col A is empty)
    # row[1] = Symbol, row[2] = ISIN, etc.
    if len(row) < 22:
        return None

    symbol = row[1]
    if not symbol or str(symbol).strip() == "":
        return None

    def safe_float(val, default=0.0):
        if val is None:
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    def safe_int(val, default=0):
        if val is None:
            return default
        try:
            return int(val)
        except (ValueError, TypeError):
            return default

    def safe_str(val, default=""):
        if val is None:
            return default
        return str(val)

    return Trade(
        symbol=safe_str(row[1]),
        isin=safe_str(row[2]),
        entry_date=safe_str(row[3]),
        exit_date=safe_str(row[4]),
        quantity=safe_float(row[5]),
        buy_value=safe_float(row[6]),
        sell_value=safe_float(row[7]),
        profit=safe_float(row[8]),
        period_of_holding=safe_int(row[9]),
        fair_market_value=safe_float(row[10]),
        taxable_profit=safe_float(row[11]),
        turnover=safe_float(row[12]),
        brokerage=safe_float(row[13]),
        exchange_charges=safe_float(row[14]),
        ipft=safe_float(row[15]),
        sebi_charges=safe_float(row[16]),
        cgst=safe_float(row[17]),
        sgst=safe_float(row[18]),
        igst=safe_float(row[19]),
        stamp_duty=safe_float(row[20]),
        stt=safe_float(row[21]),
        trade_type=trade_type,
    )


def _parse_equity_summary(wb, data: ParsedData):
    """Parse the Equity and Non Equity summary sheet."""
    ws = wb["Equity and Non Equity"]

    for row in ws.iter_rows(min_row=15, max_row=20, values_only=True):
        if not row[1]:
            continue

        label = str(row[1]).strip()
        value = row[2] if len(row) > 2 and row[2] is not None else 0.0

        try:
            value = float(value)
        except (ValueError, TypeError):
            value = 0.0

        if "Intraday" in label or "Speculative" in label:
            data.intraday_profit = value
        elif "Short Term" in label:
            data.short_term_profit = value
        elif "Long Term" in label:
            data.long_term_profit = value
        elif "Non Equity" in label:
            data.non_equity_profit = value


def _parse_charges(wb, data: ParsedData):
    """Parse the Other Debits and Credits sheet."""
    ws = wb["Other Debits and Credits"]
    current_segment = None

    for row in ws.iter_rows(min_row=13, max_row=ws.max_row, values_only=True):
        if not any(cell is not None for cell in row):
            continue

        cell_b = row[1] if len(row) > 1 else None
        if cell_b is None:
            continue

        cell_b_str = str(cell_b).strip()

        # Check if this is a segment header (only col B has value)
        if cell_b_str and all(cell is None for cell in row[2:]):
            current_segment = cell_b_str
            continue

        # Skip the "Particulars" header row
        if cell_b_str == "Particulars":
            continue

        # Parse data rows
        if current_segment and cell_b_str:
            try:
                posting_date = str(row[2]) if len(row) > 2 and row[2] else ""
                debit = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
                credit = float(row[4]) if len(row) > 4 and row[4] is not None else 0.0

                charge = Charge(
                    particulars=cell_b_str,
                    posting_date=posting_date,
                    debit=debit,
                    credit=credit,
                    segment=current_segment,
                )
                data.charges.append(charge)
            except (ValueError, TypeError, IndexError):
                continue


def _parse_dividends(wb, data: ParsedData):
    """Parse the Equity Dividends sheet."""
    ws = wb["Equity Dividends"]

    for row in ws.iter_rows(min_row=16, max_row=ws.max_row, values_only=True):
        if not any(cell is not None for cell in row):
            continue

        cell_b = row[1] if len(row) > 1 else None
        if cell_b is None:
            continue

        cell_b_str = str(cell_b).strip()

        # Stop at "Total Dividend Amount" row
        if "Total Dividend" in cell_b_str:
            if len(row) > 6 and row[6] is not None:
                try:
                    data.total_dividend = float(row[6])
                except (ValueError, TypeError):
                    pass
            break

        # Skip non-data rows
        if cell_b_str == "Symbol" or cell_b_str == "":
            continue

        # Skip info rows
        if "Dividends are credited" in cell_b_str:
            continue

        try:
            def safe_float(val, default=0.0):
                if val is None:
                    return default
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return default

            dividend = Dividend(
                symbol=cell_b_str,
                isin=str(row[2]) if len(row) > 2 and row[2] else "",
                ex_date=str(row[3]) if len(row) > 3 and row[3] else "",
                quantity=safe_float(row[4]) if len(row) > 4 else 0.0,
                dividend_per_share=safe_float(row[5]) if len(row) > 5 else 0.0,
                net_amount=safe_float(row[6]) if len(row) > 6 else 0.0,
            )
            data.dividends.append(dividend)
        except (ValueError, TypeError, IndexError):
            continue
